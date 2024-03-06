from openpyxl import Workbook
from .eemodel import (
    headers,
    EuentregoOrder,
    name_list,
    surname_list,
    random_cpf,
    random_destinatario,
)
from dataclasses import astuple
import csv
from django.conf import settings
import random
from datetime import datetime
import traceback


"""
    1. import headers
    2. import orders
    3. create file
    4. write headers
    5. write orders
    6. save to file
    7. save to stream

"""


def join_geolocation_address(address, coords):
    endereco = address["route"]
    numero = address["street_number"]
    bairro = address["sublocality"]
    municipio = address["administrative_area_level_2"]
    uf = address["administrative_area_level_1"]
    cep = address["postal_code"]

    address_with_geolocation = [
        coords[0],
        coords[1],
        endereco,
        numero,
        bairro,
        municipio,
        uf,
        cep,
    ]
    return address_with_geolocation


def create_order_from_address(address, coords):
    endereco = address["route"]
    numero = address["street_number"]
    bairro = address["sublocality"]
    municipio = address["administrative_area_level_2"]
    uf = address["administrative_area_level_1"]
    cep = address["postal_code"]

    try:
        new_order = EuentregoOrder(
            uf=uf,
            municipio=municipio,
            bairro=bairro,
            endereco=endereco,
            numero=numero,
            cep=cep,
            latitude=coords[0],
            longitude=coords[1],
            cpf_cnpj=random_cpf(),  # random cpf generation (not valid)
            destinatario=random_destinatario(),
            dt_lim=datetime.today().strftime("%d/%m/%Y"),
            carga_kg=random.randint(1, 20),
            carga_cm3=random.randint(50, 500000),
            qtd_volumes=random.randint(1, 10),
            pedido=random.randint(1000000, 9999999),
        )

        return new_order
    except Exception as e:
        print("create_ordeR_from_address")
        empty_order = EuentregoOrder(latitude=coords[0], longitude=coords[1])
        traceback.print_exc()
        return empty_order


def create_standard_ee_spreadsheet(order_list):
    wb = Workbook()
    ws = wb.active

    try:
        for h in range(1, len(headers) + 1):
            ws.cell(row=1, column=h, value=headers[h - 1])
        for x in range(2, len(order_list) + 2):
            order = order_list[x - 2]
            order_data = astuple(order)

            for y in range(1, len(order_data) + 1):
                ws.cell(row=x, column=y, value=order_data[y - 1])
    except Exception as e:
        return e

    return wb


def create_geolocation_spreadsheet(address_with_geolocation_list):
    """
    input : list/dictionary with the full addresses and their geolocation data.

    returns a openpyxl workbook with geolocation data on columns A/B and the full address on columns C-Z

    """
    wb = Workbook()
    ws = wb.active

    # TO DO : ADD HEADERS

    # TO DO : ADD ROW VALUES

    return wb


def create_geolocation_csv(filename, address_with_geolocation_list):
    try:
        file_path = f"{settings.MEDIA_ROOT}/files/{filename}.csv"
        with open(file_path, mode="w", encoding="utf-8", newline="") as csv_file:
            csv_writer = csv.writer(csv_file, delimiter=";")
            csv_writer.writerow(
                [
                    "latitude",
                    "longitude",
                    "endereço",
                    "numero",
                    "bairro",
                    "municipio",
                    "uf",
                    "país",
                    "cep",
                ]
            )
            for address in address_with_geolocation_list:
                csv_writer.writerow(address)
    except Exception as e:
        return e
    return file_path


def create_standard_ee_csv(filename, order_list):
    try:
        file_path = f"{settings.MEDIA_ROOT}/files/{filename}.csv"
        with open(file_path, mode="w", encoding="utf-8", newline="") as csv_file:
            csv_writer = csv.writer(csv_file, delimiter=";")
            csv_writer.writerow(headers)
            for order in order_list:
                csv_writer.writerow(list(astuple(order)))
    except Exception as e:
        return e
    return file_path
