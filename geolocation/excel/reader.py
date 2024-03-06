from openpyxl import load_workbook
from .eemodel import EuentregoOrder
import csv

"""
    1. load xlsx file
    2. access data from rows 2 to max_row
    3. create orders for each row
    4. return list of orders
    

"""


def load_from_standard_ee_spreadsheet(filename):
    # load
    wb = load_workbook(filename=filename)
    ws = wb.active
    # access data
    row_values = tuple(
        ws.values
    )  # ws.values is an itterable object.. isn't subscriptable
    # create orders
    order_list = []

    for row in row_values[1:]:
        rows_data_list = list(row)  # why can't i just unpack a tuple into the object
        order = EuentregoOrder(*rows_data_list)
        order_list.append(order)

    return order_list


def load_geolocation_spreadsheet(filename):
    """
    read from a .xlsx file with geolocation data on columns A (latitude) and B(longitude)

    returns list of values of latitude and longitude

    """

    wb = load_workbook(filename=filename)
    ws = wb.active
    row_values = tuple(ws.values)

    geolocation_list = []

    for row in row_values:
        ##########################VALIDATE HERE##########################
        if row != (None, None):
            geolocation_list.append(row)

    return geolocation_list


def load_address_spreadsheet(filename):
    """
    read from a .xlsx file with address strings on columns A-Z

    returns list of addresses

    """
    wb = load_workbook(filename=filename)
    ws = wb.active
    row_values = tuple(ws.values)

    address_list = []

    for row in row_values:
        address_list.append(row)

    return address_list


def load_from_standard_ee_csv(filename):

    order_list = []

    with open(filename, "r", encoding="utf8") as csv_file:

        csv_reader = csv.reader(csv_file, delimiter=";")
        line_count = 0
        for row in csv_reader:
            if line_count != 0:
                new_order = EuentregoOrder(*row)
                order_list.append(new_order)

            line_count += 1

    return order_list


def load_from_geolocation_csv(filename):
    # latitude column A, longitude column B
    geolocation_list = []

    with open(filename, "r", encoding="utf8") as csv_file:
        print("loading csv")
        csv_reader = csv.reader(csv_file, delimiter=";")
        for i, row in enumerate(csv_reader):
            geolocation_list.append((row[0], row[1]))
            print(f"loading row {i}")

    return geolocation_list
