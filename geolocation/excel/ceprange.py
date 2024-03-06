from openpyxl import load_workbook, Workbook
import sys

"""
    how to use:
    
    no prompt de comando
    
    python ceprange.py nome_arquivo_com_rotas.xlsx nome_arquivo_com_ceps_teste.xlsx

    ler da coluna D do arquivo nome_arquivo_com_ceps_teste.xlsx
    se quiser mudar a coluna, ver linha 63

    salva como resultado.xlsx
"""


def create_dictlist(filename):
    # create workbook
    wb = load_workbook(filename)

    ws = wb.active

    row_values = tuple(
        ws.values
    )  # ws.values is an itterable object.. isn't subscriptable
    #
    # extract zipcode data from workbook
    zipcode_range_list = []
    # unpacks data into a list
    for row in row_values[1:]:
        rows_data_list = list(row)
        zipcode_range_list.append(rows_data_list)

    zipcode_dict_list = []
    # iterates through the list zipcode_range and create dictionaries for each row entry
    for (
        interval
    ) in (
        zipcode_range_list
    ):  # interval is each list containing the zipcode intervals for route in the main list zipcode_range_list
        initial_zipcode = int(interval[0])
        final_zipcode = int(interval[1])
        route = int(interval[2])
        zipcode_dict = {}
        zipcode_dict["cep_inicial"] = initial_zipcode
        zipcode_dict["cep_final"] = final_zipcode
        zipcode_dict["rota"] = route
        zipcode_dict_list.append(zipcode_dict)

    return zipcode_dict_list


def load_zipcodes_to_test(filename):
    wb = load_workbook(filename)
    ws = wb.active

    row_values = tuple(ws.values)

    zipcode_list = []
    # unpacks data into a list
    for row in row_values[1:]:
        rows_data_list = int(row[3])  # loads zipcodes from column D
        zipcode_list.append(rows_data_list)

    return zipcode_list


def check_route(zipcode, zipcode_dict_list):
    routes_matched = []

    for d in zipcode_dict_list:
        if zipcode >= d["cep_inicial"] and zipcode <= d["cep_final"]:
            routes_matched.append([zipcode, d["rota"], zipcode_dict_list.index(d) + 2])

    if routes_matched:
        routes = [rota[1] for rota in routes_matched]
        lines = [linha[2] for linha in routes_matched]
        return f"CEP {zipcode} está na rota {routes} na linha {lines}"

    else:
        return f"Rota não encontrada pro cep {zipcode}"


def save_to_xlsx(result):
    wb = Workbook()
    ws = wb.active

    for r in range(1, len(result) + 1):
        ws.cell(row=r, column=1, value=result[r - 1])

    wb.save("resultado.xlsx")

    print("aqruivo salvo")


if __name__ == "__main__":

    to_check = []

    filename1 = sys.argv[1]
    filename2 = sys.argv[2]

    zipcode_dict_list = create_dictlist(filename=filename1)
    zipcodes_to_check = load_zipcodes_to_test(filename=filename2)

    result = []

    for zipcode in zipcodes_to_check:
        result.append(check_route(zipcode, zipcode_dict_list))

    save_to_xlsx(result)
